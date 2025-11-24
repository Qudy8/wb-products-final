"""Тестовый скрипт для проверки Excel файла"""
import os
from openpyxl import load_workbook


def test_excel_file():
    """Проверяем структуру Excel файла"""

    # Находим все Excel файлы в user_files
    user_files_dir = "user_files"

    if not os.path.exists(user_files_dir):
        print(f"❌ Папка {user_files_dir} не найдена")
        return

    excel_files = [f for f in os.listdir(user_files_dir) if f.endswith(('.xlsx', '.xls'))]

    if not excel_files:
        print(f"❌ Excel файлы не найдены в {user_files_dir}")
        return

    print(f"Найдено Excel файлов: {len(excel_files)}")

    for filename in excel_files:
        filepath = os.path.join(user_files_dir, filename)
        print(f"\n{'='*80}")
        print(f"Файл: {filename}")
        print(f"{'='*80}")

        try:
            workbook = load_workbook(filepath, read_only=True, data_only=True)
            sheet = workbook.active

            print(f"Активный лист: {sheet.title}")
            print(f"Размер: {sheet.max_row} строк × {sheet.max_column} столбцов")

            print(f"\n{'='*80}")
            print("Первые 10 строк:")
            print(f"{'='*80}")

            for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if idx > 10:
                    break

                # Показываем первые 5 столбцов
                row_data = row[:5] if row else []
                print(f"Строка {idx}: {row_data}")

            print(f"\n{'='*80}")
            print("Анализ данных:")
            print(f"{'='*80}")

            # Проверяем заголовки (строка 1)
            headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if headers:
                print(f"Заголовки (строка 1):")
                for idx, header in enumerate(headers[:5], start=1):
                    col_letter = chr(64 + idx)  # A=65, B=66, etc
                    print(f"  Столбец {col_letter}: '{header}'")

            # Считаем непустые строки
            non_empty_rows = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and any(cell for cell in row[:3]):  # Проверяем первые 3 столбца
                    non_empty_rows += 1

            print(f"\nНепустых строк с данными (начиная со строки 2): {non_empty_rows}")

            # Показываем пример данных
            print(f"\nПример данных (строка 2):")
            example = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True), None)
            if example:
                print(f"  A (Категория): '{example[0] if len(example) > 0 else ''}'")
                print(f"  B (Предмет): '{example[1] if len(example) > 1 else ''}'")
                print(f"  C (Комиссия): '{example[2] if len(example) > 2 else ''}'")
            else:
                print("  ❌ Нет данных в строке 2")

            workbook.close()

        except Exception as e:
            print(f"❌ Ошибка при чтении файла: {e}")
            import traceback
            traceback.print_exc()


if __name__ == "__main__":
    test_excel_file()
