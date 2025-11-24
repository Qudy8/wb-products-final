"""Отладка загрузки Excel"""
from openpyxl import load_workbook

excel_path = "user_files/701912845_сomission.xlsx"

print("Отладка загрузки Excel файла...\n")

try:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook.active

    print(f"✅ Файл открыт")
    print(f"✅ Активный лист: {sheet.title}")
    print(f"✅ Max row: {sheet.max_row}")
    print(f"✅ Max column: {sheet.max_column}")

    data = {}
    processed = 0
    skipped = 0

    print(f"\nОбработка строк...")

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row_idx <= 5:  # Отладка первых строк
            print(f"\nСтрока {row_idx}:")
            print(f"  Длина row: {len(row) if row else 0}")
            print(f"  row: {row}")

        if not row or len(row) < 2:
            skipped += 1
            if row_idx <= 5:
                print(f"  ❌ ПРОПУЩЕНА (not row or len < 2)")
            continue

        category = row[0] if len(row) > 0 else None
        subject = row[1] if len(row) > 1 else None
        commission_wb = row[2] if len(row) > 2 else None
        commission_fbs = row[3] if len(row) > 3 else None

        if row_idx <= 5:
            print(f"  category: '{category}'")
            print(f"  subject: '{subject}'")
            print(f"  commission_wb: '{commission_wb}'")
            print(f"  commission_fbs: '{commission_fbs}'")

        if not subject:
            skipped += 1
            if row_idx <= 5:
                print(f"  ❌ ПРОПУЩЕНА (subject пустой)")
            continue

        subject_key = str(subject).strip().lower()

        if row_idx <= 5:
            print(f"  subject_key: '{subject_key}'")

        data[subject_key] = {
            'category': str(category).strip() if category else '',
            'subject': str(subject).strip() if subject else '',
            'commission_wb': str(commission_wb).strip() if commission_wb else '',
            'commission_fbs': str(commission_fbs).strip() if commission_fbs else ''
        }

        processed += 1

        if row_idx <= 5:
            print(f"  ✅ ДОБАВЛЕНА в data")

    print(f"\n{'='*80}")
    print(f"Итого:")
    print(f"  Обработано: {processed}")
    print(f"  Пропущено: {skipped}")
    print(f"  Записей в data: {len(data)}")

    # Показываем первые 3 записи
    print(f"\nПервые 3 записи в data:")
    for idx, (key, value) in enumerate(list(data.items())[:3], 1):
        print(f"\n{idx}. '{key}'")
        print(f"   {value}")

    workbook.close()

except Exception as e:
    print(f"❌ Ошибка: {e}")
    import traceback
    traceback.print_exc()
