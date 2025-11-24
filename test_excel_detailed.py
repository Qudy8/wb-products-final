"""Детальный анализ Excel файла"""
import os
from openpyxl import load_workbook


def analyze_excel():
    """Детальный анализ структуры Excel"""

    filepath = "user_files/701912845_сomission.xlsx"

    if not os.path.exists(filepath):
        print("❌ Файл не найден")
        return

    print(f"Анализ файла: {filepath}\n")

    workbook = load_workbook(filepath, read_only=False, data_only=True)
    sheet = workbook.active

    print(f"Активный лист: {sheet.title}")
    print(f"Диапазон используемых ячеек: {sheet.dimensions}")
    print(f"Max row (по данным openpyxl): {sheet.max_row}")
    print(f"Max column (по данным openpyxl): {sheet.max_column}")

    print(f"\n{'='*80}")
    print("Первые 20 строк (все столбцы A-E):")
    print(f"{'='*80}\n")

    for row_idx in range(1, min(21, sheet.max_row + 1)):
        print(f"Строка {row_idx}:")
        row_data = []
        for col_idx in range(1, 6):  # Столбцы A-E (1-5)
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            col_letter = chr(64 + col_idx)
            print(f"  {col_letter}: '{cell_value}'")
            row_data.append(cell_value)
        print()

    print(f"{'='*80}")
    print("Проверка наличия данных по столбцам:")
    print(f"{'='*80}\n")

    for col_idx in range(1, 6):
        col_letter = chr(64 + col_idx)
        non_empty = 0
        for row_idx in range(1, sheet.max_row + 1):
            if sheet.cell(row=row_idx, column=col_idx).value:
                non_empty += 1
        print(f"Столбец {col_letter}: {non_empty} непустых ячеек")

    # Проверяем merged cells (объединенные ячейки)
    print(f"\n{'='*80}")
    print("Объединенные ячейки:")
    print(f"{'='*80}\n")

    if sheet.merged_cells:
        for merged_range in sheet.merged_cells.ranges:
            print(f"  {merged_range}")
    else:
        print("  Нет объединенных ячеек")

    workbook.close()


if __name__ == "__main__":
    analyze_excel()
