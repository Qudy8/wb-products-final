"""Модуль для работы с Excel файлами"""
from openpyxl import load_workbook
from typing import Dict, Optional


class ExcelHelper:
    """Класс для работы с Excel файлами с данными о категориях и комиссиях"""

    def __init__(self, file_path: str):
        """
        Инициализация

        Args:
            file_path: путь к Excel файлу
        """
        self.file_path = file_path
        self.data_cache = None

    def load_data(self) -> Dict[str, Dict[str, str]]:
        """
        Загружает данные из Excel файла

        Структура файла:
        - Столбец A: Категория
        - Столбец B: Предмет
        - Столбец C: Комиссия WB
        - Столбец D: Комиссия FBS (везу на склад WB)
        - Столбец E: Комиссия самостоятельной доставки

        Returns:
            Словарь вида {предмет: {category, commission_wb, commission_fbs, commission_self}}
        """
        if self.data_cache is not None:
            return self.data_cache

        try:
            # Открываем файл без read_only для правильного определения размера
            workbook = load_workbook(self.file_path, data_only=True)
            sheet = workbook.active

            data = {}

            # Определяем реальный диапазон данных
            # Используем dimensions для получения реального диапазона
            max_row = sheet.max_row

            # Если max_row = 1, пытаемся найти реальный размер
            if max_row == 1:
                # Перебираем строки до первой пустой
                for row_num in range(1, 10000):  # Лимит 10000 строк
                    cell_value = sheet.cell(row=row_num, column=1).value
                    if cell_value is None and row_num > 2:
                        max_row = row_num - 1
                        break
                    if row_num == 9999:
                        max_row = 9999
                        break

            # Начинаем со второй строки (пропускаем заголовок)
            for row_idx in range(2, max_row + 1):
                # Читаем ячейки напрямую
                category = sheet.cell(row=row_idx, column=1).value  # Столбец A
                subject = sheet.cell(row=row_idx, column=2).value   # Столбец B
                commission_wb = sheet.cell(row=row_idx, column=3).value   # Столбец C
                commission_fbs = sheet.cell(row=row_idx, column=4).value  # Столбец D (FBS)
                commission_self = sheet.cell(row=row_idx, column=5).value # Столбец E

                # Пропускаем пустые строки
                if not subject:
                    continue

                # Сохраняем данные, используя предмет как ключ
                # Приводим к нижнему регистру и убираем лишние пробелы для сравнения
                subject_key = str(subject).strip().lower()

                data[subject_key] = {
                    'category': str(category).strip() if category else '',
                    'subject': str(subject).strip() if subject else '',
                    'commission_wb': str(commission_wb).strip() if commission_wb else '',
                    'commission_fbs': str(commission_fbs).strip() if commission_fbs else '',
                    'commission_self': str(commission_self).strip() if commission_self else ''
                }

            workbook.close()
            self.data_cache = data
            return data

        except Exception as e:
            print(f"Ошибка при чтении Excel файла: {e}")
            import traceback
            traceback.print_exc()
            return {}

    def _normalize_text(self, text: str) -> str:
        """
        Нормализует текст для сравнения:
        - нижний регистр
        - удаление лишних пробелов
        - приведение к единственному числу
        """
        text = str(text).strip().lower()

        # Нормализуем каждое слово отдельно
        words = text.split()
        normalized_words = []

        for word in words:
            # Убираем окончания множественного числа
            # "коврики" -> "коврик", "лонгсливы" -> "лонгслив"
            if len(word) > 4:  # Обрабатываем только слова длиннее 4 букв
                if word.endswith('ики'):
                    word = word[:-1]  # "коврики" -> "коврик"
                elif word.endswith('и') and not word.endswith('ии'):
                    word = word[:-1]  # "вибраторы" -> убираем но проверяем чтобы не "линии"
                elif word.endswith('ы'):
                    word = word[:-1]  # "домкраты" -> "домкрат"

            normalized_words.append(word)

        return ' '.join(normalized_words)

    def find_by_subject(self, entity: str) -> Optional[Dict[str, str]]:
        """
        Ищет информацию по предмету (entity)

        Args:
            entity: название предмета из WB API

        Returns:
            Словарь с данными {category, subject, commission} или None
        """
        data = self.load_data()

        if not data:
            return None

        # Приводим к нижнему регистру для поиска
        entity_key = str(entity).strip().lower()

        # Сначала пробуем точное совпадение
        if entity_key in data:
            return data.get(entity_key)

        # Если не нашли, пробуем нормализованное совпадение
        entity_normalized = self._normalize_text(entity)

        for key, value in data.items():
            key_normalized = self._normalize_text(key)

            # Проверяем нормализованное совпадение
            if entity_normalized == key_normalized:
                return value

            # Проверяем частичное совпадение (один содержит другой)
            if entity_normalized in key_normalized or key_normalized in entity_normalized:
                # Если длины похожи (разница не более 3 символов), считаем совпадением
                if abs(len(entity_normalized) - len(key_normalized)) <= 3:
                    return value

        return None

    def get_stats(self) -> Dict[str, int]:
        """
        Возвращает статистику по загруженным данным

        Returns:
            Словарь со статистикой
        """
        data = self.load_data()

        categories = set()
        for item in data.values():
            if item['category']:
                categories.add(item['category'])

        return {
            'total_subjects': len(data),
            'total_categories': len(categories)
        }
